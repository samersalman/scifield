"""Smoke + round-trip tests for the `scifield epistemic` Typer sub-app (V1-S07).

These tests cover the CLI plumbing only — sampling correctness lives in
``test_epistemic_sampling.py``, extractor correctness in
``test_epistemic_extract.py``, etc. We patch the heavy pieces
(`stratified_sample`, `extract_one`) and exercise the four commands end
to end with tmp_path fixtures.
"""

from __future__ import annotations

import sys
import types
from pathlib import Path
from typing import Any

import duckdb
import pandas as pd
import pytest
from omegaconf import DictConfig, OmegaConf
from openpyxl import Workbook
from typer.testing import CliRunner

import scifield.cli as cli_mod
from scifield.cli import app
from scifield.epistemic.labeling import LABELS_HEADER
from scifield.epistemic.schema import EpistemicExtraction, EpistemicLabel

runner = CliRunner()


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _make_synth_duckdb(path: Path) -> None:
    """Plant a tiny ``papers`` table sufficient for ``ensure_papers_distinct_view``."""
    con = duckdb.connect(str(path))
    try:
        con.execute(
            """
            CREATE TABLE papers (
                pmid BIGINT,
                journal_slug VARCHAR,
                title VARCHAR,
                abstract VARCHAR,
                year INTEGER,
                fetched_at VARCHAR
            )
            """
        )
        rows = []
        for i in range(5):
            rows.append(
                (
                    1_000_000 + i,
                    "ann_surg",
                    f"title-{i}",
                    "A" * 80 + f" synthetic abstract {i}",
                    2015,
                    "2026-01-01T00:00:00Z",
                )
            )
        con.executemany(
            "INSERT INTO papers VALUES (?, ?, ?, ?, ?, ?)",
            rows,
        )
    finally:
        con.close()


def _make_synth_topics(path: Path) -> None:
    df = pd.DataFrame(
        {
            "pmid": [1_000_000 + i for i in range(5)],
            "topic_id": [0, 1, 2, 3, 4],
            "is_noise": [False] * 5,
        }
    )
    df.to_parquet(path, index=False)


def _make_synth_sample_parquet(path: Path, n_rows: int = 3) -> pd.DataFrame:
    df = pd.DataFrame(
        {
            "pmid": [1_000_000 + i for i in range(n_rows)],
            "journal": ["ann_surg"] * n_rows,
            "year": [2015] * n_rows,
            "era": ["2010-2019"] * n_rows,
            "topic_id": pd.array([0, 1, 2][:n_rows], dtype="Int64"),
            "title": [f"title-{i}" for i in range(n_rows)],
            "abstract": ["A" * 80 + f" synthetic abstract {i}" for i in range(n_rows)],
        }
    )
    df.to_parquet(path, index=False)
    return df


def _build_synth_cfg(tmp_path: Path) -> DictConfig:
    duckdb_path = tmp_path / "papers.duckdb"
    topics_parquet = tmp_path / "topics.parquet"
    sample_path = tmp_path / "handlabel_sample.parquet"
    handlabel_parquet = tmp_path / "epistemic_handlabel.parquet"
    pilot_path = tmp_path / "epistemic_pilot.parquet"
    pilot_failed_path = tmp_path / "epistemic_pilot_failed.parquet"
    labels_xlsx_dir = tmp_path
    extract_out = tmp_path / "epistemic_extracted.parquet"
    extract_failed = tmp_path / "epistemic_failed.parquet"
    return OmegaConf.create(
        {
            "input": {
                "duckdb_path": str(duckdb_path),
                "topics_parquet": str(topics_parquet),
            },
            "output": {
                "sample_path": str(sample_path),
                "handlabel_parquet": str(handlabel_parquet),
                "pilot_path": str(pilot_path),
                "pilot_failed_path": str(pilot_failed_path),
                "labels_xlsx_dir": str(labels_xlsx_dir),
            },
            "sampling": {
                "n_sample": 3,
                "seed": 42,
                "eras": ["pre2000", "2000-2009", "2010-2019", "2020+"],
                "topic_coverage_min": 1,
            },
            "pilot": {
                "n_pilot": 3,
                "claude_cmd": ["claude", "--print"],
            },
            "model": {"id": "claude-via-claude-code"},
            "prompt": {"version": "v0.1"},
            "extract_batch": {
                "out_path": str(extract_out),
                "failed_path": str(extract_failed),
                "concurrency": 1,
                "chunk_size": 500,
                "preregistration_url": "https://doi.org/10.17605/OSF.IO/8ZJHD",
            },
        }
    )


# ---------------------------------------------------------------------------
# Help
# ---------------------------------------------------------------------------


def test_help_works() -> None:
    result = runner.invoke(app, ["epistemic", "--help"])
    assert result.exit_code == 0, result.stdout
    for cmd in ("sample", "export-labels", "import-labels", "pilot"):
        assert cmd in result.stdout, f"missing {cmd!r} in help output"


# ---------------------------------------------------------------------------
# sample
# ---------------------------------------------------------------------------


def test_sample_smoke(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    """End-to-end the `sample` command with `stratified_sample` patched.

    The CLI is responsible for: loading config, ensuring the
    `papers_distinct` view exists, calling the sampler, writing parquet,
    and recording the run sidecar. We stub `stratified_sample` so this
    test exercises only the CLI plumbing.
    """
    cfg = _build_synth_cfg(tmp_path)
    _make_synth_duckdb(Path(str(cfg.input.duckdb_path)))
    _make_synth_topics(Path(str(cfg.input.topics_parquet)))

    monkeypatch.setattr(cli_mod, "_load_epistemic_config", lambda name: cfg)

    fake_df = pd.DataFrame(
        {
            "pmid": [1_000_000, 1_000_001, 1_000_002],
            "journal": ["ann_surg", "ann_surg", "ann_surg"],
            "year": [2015, 2015, 2015],
            "era": ["2010-2019", "2010-2019", "2010-2019"],
            "topic_id": pd.array([0, 1, 2], dtype="Int64"),
            "title": ["t0", "t1", "t2"],
            "abstract": ["a0", "a1", "a2"],
        }
    )

    def _fake_stratified_sample(con: Any, sampling_cfg: Any) -> pd.DataFrame:
        return fake_df

    # The CLI body does `from scifield.epistemic.sampling import ... stratified_sample`,
    # so we patch the symbol on that module.
    import scifield.epistemic.sampling as sampling_mod

    monkeypatch.setattr(sampling_mod, "stratified_sample", _fake_stratified_sample)

    result = runner.invoke(
        app,
        ["epistemic", "sample", "--n-sample", "3", "--seed", "42"],
    )
    assert result.exit_code == 0, result.stdout

    sample_path = Path(str(cfg.output.sample_path))
    assert sample_path.exists(), "sample parquet was not written"
    sidecar = Path(str(sample_path) + ".run.json")
    assert sidecar.exists(), "run.json sidecar was not written"
    written = pd.read_parquet(sample_path)
    assert len(written) == 3
    assert set(written.columns) >= {"pmid", "journal", "year", "era", "title", "abstract"}


def test_sample_errors_on_missing_db(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    cfg = _build_synth_cfg(tmp_path)
    # Do NOT create the duckdb; only topics.parquet.
    _make_synth_topics(Path(str(cfg.input.topics_parquet)))
    monkeypatch.setattr(cli_mod, "_load_epistemic_config", lambda name: cfg)

    result = runner.invoke(app, ["epistemic", "sample"])
    assert result.exit_code == 1
    assert "papers DuckDB not found" in result.stdout


# ---------------------------------------------------------------------------
# export-labels
# ---------------------------------------------------------------------------


def test_export_labels_smoke(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    cfg = _build_synth_cfg(tmp_path)
    sample_path = Path(str(cfg.output.sample_path))
    sample_path.parent.mkdir(parents=True, exist_ok=True)
    _make_synth_sample_parquet(sample_path, n_rows=3)

    monkeypatch.setattr(cli_mod, "_load_epistemic_config", lambda name: cfg)

    result = runner.invoke(app, ["epistemic", "export-labels", "--rater", "test"])
    assert result.exit_code == 0, result.stdout

    expected = Path(str(cfg.output.labels_xlsx_dir)) / "labels_test.xlsx"
    assert expected.exists(), f"expected xlsx at {expected} was not written"
    assert f"wrote {expected}" in result.stdout


def test_export_labels_errors_on_missing_sample(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    cfg = _build_synth_cfg(tmp_path)
    # Do NOT create the sample parquet.
    monkeypatch.setattr(cli_mod, "_load_epistemic_config", lambda name: cfg)

    result = runner.invoke(app, ["epistemic", "export-labels", "--rater", "test"])
    assert result.exit_code == 1
    assert "sample parquet not found" in result.stdout


# ---------------------------------------------------------------------------
# import-labels
# ---------------------------------------------------------------------------


def _build_labels_xlsx(path: Path, pmid: int = 1_000_000) -> None:
    """Fabricate a one-row labeled workbook against ``LABELS_HEADER``."""
    wb = Workbook()
    # First sheet becomes Instructions; we just need a Labels sheet.
    wb.active.title = "Instructions"
    ws = wb.create_sheet("Labels")
    ws.append(list(LABELS_HEADER))
    ws.append(
        [
            pmid,
            "ann_surg",
            2015,
            "title-x",
            "abstract-x" * 8,
            "RCT",  # study_design
            120,  # sample_size
            "TRUE",  # has_control
            "positive",  # effect_direction
            "TRUE",  # statistical_claim_present
            "FALSE",  # coi_disclosed_in_abstract
        ]
    )
    wb.save(path)


def test_import_labels_round_trip(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    cfg = _build_synth_cfg(tmp_path)
    xlsx_path = tmp_path / "labels_test.xlsx"
    _build_labels_xlsx(xlsx_path, pmid=1_000_000)

    monkeypatch.setattr(cli_mod, "_load_epistemic_config", lambda name: cfg)

    result = runner.invoke(
        app,
        ["epistemic", "import-labels", "--rater", "test", "--file", str(xlsx_path)],
    )
    assert result.exit_code == 0, result.stdout

    parquet_out = Path(str(cfg.output.handlabel_parquet))
    assert parquet_out.exists()
    df = pd.read_parquet(parquet_out)
    # 6 long-form rows per labeled abstract (one per RATER_FILL field).
    assert len(df) == 6
    assert set(df["pmid"].unique()) == {1_000_000}
    assert set(df["rater"].unique()) == {"test"}


def test_import_labels_exits_1_on_errors(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    cfg = _build_synth_cfg(tmp_path)
    monkeypatch.setattr(cli_mod, "_load_epistemic_config", lambda name: cfg)

    # Workbook with one row that has a bad enum -> import collects an error.
    xlsx_path = tmp_path / "labels_bad.xlsx"
    wb = Workbook()
    wb.active.title = "Instructions"
    ws = wb.create_sheet("Labels")
    ws.append(list(LABELS_HEADER))
    ws.append(
        [
            1_000_000,
            "ann_surg",
            2015,
            "t",
            "abstract" * 8,
            "NOT_A_VALID_DESIGN",  # bad enum
            120,
            "TRUE",
            "positive",
            "TRUE",
            "FALSE",
        ]
    )
    wb.save(xlsx_path)

    result = runner.invoke(
        app,
        ["epistemic", "import-labels", "--rater", "test", "--file", str(xlsx_path)],
    )
    assert result.exit_code == 1, result.stdout
    assert "n_errors=1" in result.stdout


# ---------------------------------------------------------------------------
# pilot
# ---------------------------------------------------------------------------


def _install_fake_pilot_module(
    monkeypatch: pytest.MonkeyPatch,
    canned: EpistemicExtraction,
) -> None:
    """Install a minimal `scifield.epistemic.pilot` module for the CLI to import.

    Batch 4A owns the real implementation; this stub is just enough for
    the CLI command body to do its lazy import and run.
    """
    from dataclasses import dataclass

    fake = types.ModuleType("scifield.epistemic.pilot")

    @dataclass(frozen=True)
    class PilotConfig:  # noqa: D401 - stub
        sample_path: Path
        pilot_path: Path
        pilot_failed_path: Path
        n_pilot: int
        extract_cfg: Any

    def run_pilot(cfg: PilotConfig, progress: Any = None) -> dict:
        df = pd.read_parquet(cfg.sample_path)
        n = min(int(cfg.n_pilot), len(df))
        ok = 0
        for i in range(n):
            if progress is not None:
                progress(i + 1, n, "ok")
            ok += 1
        # Write a minimal pilot parquet so the assertion below has
        # something to find.
        pd.DataFrame(
            {
                "pmid": df["pmid"].iloc[:n].tolist(),
                "model_id": [canned.model_id] * n,
                "prompt_version": [canned.prompt_version] * n,
            }
        ).to_parquet(cfg.pilot_path, index=False)
        return {
            "n_ok": ok,
            "n_failed": 0,
            "pilot_path": str(cfg.pilot_path),
            "pilot_failed_path": str(cfg.pilot_failed_path),
        }

    fake.PilotConfig = PilotConfig  # type: ignore[attr-defined]
    fake.run_pilot = run_pilot  # type: ignore[attr-defined]
    monkeypatch.setitem(sys.modules, "scifield.epistemic.pilot", fake)


def test_pilot_runs_with_mocked_extract(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    cfg = _build_synth_cfg(tmp_path)
    sample_path = Path(str(cfg.output.sample_path))
    sample_path.parent.mkdir(parents=True, exist_ok=True)
    _make_synth_sample_parquet(sample_path, n_rows=3)

    canned = EpistemicExtraction(
        pmid=1_000_000,
        label=EpistemicLabel(
            study_design="RCT",
            sample_size=120,
            has_control=True,
            effect_direction="positive",
            statistical_claim_present=True,
            coi_disclosed_in_abstract=False,
        ),
        model_id="claude-via-claude-code",
        prompt_version="v0.1",
        raw_response='{"study_design":"RCT"}',
    )

    _install_fake_pilot_module(monkeypatch, canned)
    monkeypatch.setattr(cli_mod, "_load_epistemic_config", lambda name: cfg)

    result = runner.invoke(app, ["epistemic", "pilot", "--n", "3"])
    assert result.exit_code == 0, result.stdout

    pilot_path = Path(str(cfg.output.pilot_path))
    assert pilot_path.exists(), "pilot parquet was not written"
    df = pd.read_parquet(pilot_path)
    assert len(df) == 3
    assert "pilot done" in result.stdout


# ---------------------------------------------------------------------------
# extract-batch
# ---------------------------------------------------------------------------


def _seed_extracted_parquet(path: Path, pmid: int) -> None:
    """Plant a one-row success parquet matching PILOT_SUCCESS_SCHEMA."""
    import pyarrow as pa
    import pyarrow.parquet as pq

    from scifield.epistemic.pilot import PILOT_SUCCESS_SCHEMA

    row = {
        "pmid": pmid,
        "study_design": "RCT",
        "sample_size": 100,
        "has_control": True,
        "effect_direction": "positive",
        "statistical_claim_present": True,
        "coi_disclosed_in_abstract": False,
        "model_id": "claude-via-claude-code",
        "prompt_version": "v0.1",
        "raw_response": '{"study_design":"RCT"}',
        "extracted_at": "2026-01-01T00:00:00+00:00",
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    table = pa.Table.from_pylist([row], schema=PILOT_SUCCESS_SCHEMA)
    pq.write_table(table, path)


def test_cli_extract_batch_status_smoke(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    """`extract-batch --status` reports n_total / n_done / n_failed / n_remaining."""
    cfg = _build_synth_cfg(tmp_path)
    _make_synth_duckdb(Path(str(cfg.input.duckdb_path)))
    # Plant one already-done row so the status counters split non-trivially.
    out_path = Path(str(cfg.extract_batch.out_path))
    _seed_extracted_parquet(out_path, pmid=1_000_000)

    monkeypatch.setattr(cli_mod, "_load_epistemic_config", lambda name: cfg)

    result = runner.invoke(app, ["epistemic", "extract-batch", "--status"])
    assert result.exit_code == 0, result.stdout
    assert "n_total=" in result.stdout
    assert "n_done=" in result.stdout
    assert "n_failed=" in result.stdout
    assert "n_remaining=" in result.stdout
    # We planted 5 rows total and pre-seeded 1, so the headline counts must
    # appear with these exact integer values.
    assert "n_total=5" in result.stdout, result.stdout
    assert "n_done=1" in result.stdout, result.stdout


def test_cli_extract_batch_submit_smoke(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    """`extract-batch --submit` runs end-to-end with subprocess + executor patched."""
    import json as _json
    from unittest.mock import MagicMock, patch

    from scifield.epistemic.batch import concurrent as _bc

    cfg = _build_synth_cfg(tmp_path)
    _make_synth_duckdb(Path(str(cfg.input.duckdb_path)))

    monkeypatch.setattr(cli_mod, "_load_epistemic_config", lambda name: cfg)

    # Inline executor + subprocess router (mirrors tests/test_epistemic_batch.py).
    class _InlineFuture:
        def __init__(self, value: Any = None) -> None:
            self._value = value

        def result(self) -> Any:
            return self._value

    class _InlineExecutor:
        def __init__(self, max_workers: int = 1, **_: Any) -> None:
            self.max_workers = max_workers

        def __enter__(self) -> _InlineExecutor:
            return self

        def __exit__(self, *exc_info: Any) -> None:
            return None

        def submit(self, fn: Any, *args: Any, **kwargs: Any) -> _InlineFuture:
            return _InlineFuture(value=fn(*args, **kwargs))

    def _inline_as_completed(future_to_pmid: dict[_InlineFuture, int]) -> list[_InlineFuture]:
        return list(future_to_pmid.keys())

    valid_payload = _json.dumps(
        {
            "study_design": "RCT",
            "sample_size": 240,
            "has_control": True,
            "effect_direction": "positive",
            "statistical_claim_present": True,
            "coi_disclosed_in_abstract": False,
        }
    )

    import subprocess as _subprocess

    real_run = _subprocess.run

    def _router(*args: Any, **kwargs: Any) -> Any:
        argv = args[0] if args else kwargs.get("args")
        if isinstance(argv, list | tuple) and len(argv) > 0 and argv[0] == "claude":
            proc = MagicMock(spec=_subprocess.CompletedProcess)
            proc.stdout = valid_payload
            proc.stderr = ""
            proc.returncode = 0
            return proc
        return real_run(*args, **kwargs)

    with (
        patch.object(_bc.futures, "ProcessPoolExecutor", new=_InlineExecutor),
        patch.object(_bc.futures, "as_completed", new=_inline_as_completed),
        patch("scifield.epistemic.extract.subprocess.run", side_effect=_router),
    ):
        result = runner.invoke(
            app,
            [
                "epistemic",
                "extract-batch",
                "--submit",
                "--concurrency",
                "1",
                "--limit",
                "2",
            ],
        )

    assert result.exit_code == 0, result.stdout
    assert "extract-batch done" in result.stdout
    out_path = Path(str(cfg.extract_batch.out_path))
    assert out_path.exists(), "extracted parquet was not created"
    df = pd.read_parquet(out_path)
    assert len(df) == 2


def test_cli_extract_batch_requires_exactly_one_mode(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    """Neither flag (or both) -> exit 2 with helpful message."""
    cfg = _build_synth_cfg(tmp_path)
    monkeypatch.setattr(cli_mod, "_load_epistemic_config", lambda name: cfg)

    result_none = runner.invoke(app, ["epistemic", "extract-batch"])
    assert result_none.exit_code == 2, result_none.stdout

    result_both = runner.invoke(app, ["epistemic", "extract-batch", "--submit", "--status"])
    assert result_both.exit_code == 2, result_both.stdout


# ---------------------------------------------------------------------------
# arbitrate-export / arbitrate-import
# ---------------------------------------------------------------------------


RATER_A_TAG = "samer"
RATER_B_TAG = "partner"


def _build_two_rater_handlabel_parquet(out_path: Path) -> None:
    """Plant a synthetic long-form parquet with one agreement + one disagreement pmid."""
    base_fields = {
        "study_design": "RCT",
        "sample_size": "240",
        "has_control": "true",
        "effect_direction": "positive",
        "statistical_claim_present": "true",
        "coi_disclosed_in_abstract": "false",
    }

    def _rows_for(pmid: int, rater: str, overrides: dict[str, str] | None = None) -> list[dict]:
        merged = {**base_fields, **(overrides or {})}
        return [
            {
                "pmid": pmid,
                "rater": rater,
                "field": field,
                "value": value,
                "imported_at": "2026-05-23T00:00:00+00:00",
            }
            for field, value in merged.items()
        ]

    rows: list[dict] = []
    # pmid 2000 — full agreement.
    rows.extend(_rows_for(2000, RATER_A_TAG))
    rows.extend(_rows_for(2000, RATER_B_TAG))
    # pmid 2001 — study_design disagreement.
    rows.extend(_rows_for(2001, RATER_A_TAG))
    rows.extend(_rows_for(2001, RATER_B_TAG, overrides={"study_design": "cohort"}))

    df = pd.DataFrame(rows, columns=["pmid", "rater", "field", "value", "imported_at"])
    out_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_parquet(out_path, index=False)


def test_cli_arbitrate_export_smoke(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    """`arbitrate-export` writes an xlsx workbook with the 5-column Arbitration sheet."""
    cfg = _build_synth_cfg(tmp_path)
    monkeypatch.setattr(cli_mod, "_load_epistemic_config", lambda name: cfg)

    handlabel_path = Path(str(cfg.output.handlabel_parquet))
    _build_two_rater_handlabel_parquet(handlabel_path)

    out_path = tmp_path / "arbitration.xlsx"
    result = runner.invoke(
        app,
        [
            "epistemic",
            "arbitrate-export",
            "--rater-a",
            RATER_A_TAG,
            "--rater-b",
            RATER_B_TAG,
            "--out",
            str(out_path),
        ],
    )
    assert result.exit_code == 0, result.stdout
    assert out_path.exists(), "arbitration xlsx was not written"
    # We planted one (pmid 2001) study_design disagreement.
    assert "n_disagreements=1" in result.stdout


def test_cli_arbitrate_import_smoke(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    """`arbitrate-import` round-trips agreements + filled-in arbitration to final parquet."""
    from openpyxl import load_workbook

    cfg = _build_synth_cfg(tmp_path)
    monkeypatch.setattr(cli_mod, "_load_epistemic_config", lambda name: cfg)

    handlabel_path = Path(str(cfg.output.handlabel_parquet))
    _build_two_rater_handlabel_parquet(handlabel_path)

    arb_path = tmp_path / "arbitration.xlsx"
    export_result = runner.invoke(
        app,
        [
            "epistemic",
            "arbitrate-export",
            "--rater-a",
            RATER_A_TAG,
            "--rater-b",
            RATER_B_TAG,
            "--out",
            str(arb_path),
        ],
    )
    assert export_result.exit_code == 0, export_result.stdout

    # Fill the 'final' column for the one disagreement.
    wb = load_workbook(arb_path)
    ws = wb["Arbitration"]
    header_row = [c.value for c in ws[1]]
    final_col_idx = header_row.index("final") + 1
    pmid_col_idx = header_row.index("pmid") + 1
    field_col_idx = header_row.index("field") + 1
    for row in ws.iter_rows(min_row=2):
        pmid_cell = row[pmid_col_idx - 1].value
        field_cell = row[field_col_idx - 1].value
        if pmid_cell is None or field_cell is None:
            continue
        if int(pmid_cell) == 2001 and str(field_cell) == "study_design":
            ws.cell(row=row[0].row, column=final_col_idx, value="RCT")
    wb.save(arb_path)

    final_path = tmp_path / "epistemic_handlabel_final.parquet"
    result = runner.invoke(
        app,
        [
            "epistemic",
            "arbitrate-import",
            "--rater-a",
            RATER_A_TAG,
            "--rater-b",
            RATER_B_TAG,
            "--arbitration",
            str(arb_path),
            "--out",
            str(final_path),
        ],
    )
    assert result.exit_code == 0, result.stdout
    assert final_path.exists()
    df = pd.read_parquet(final_path)
    assert set(df["pmid"].unique()) == {2000, 2001}
    assert "n_pmids=2" in result.stdout
