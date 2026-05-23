"""Tests for V1-S07 Excel labeling round-trip (Batch 2D).

Plants a 5-row synthetic stratified sample, drives ``export_to_xlsx`` to
build the workbook, then programmatically fills in label cells via
:mod:`openpyxl` and runs ``import_from_xlsx`` against the result. The
shape-of-the-output assertions live here (long-form parquet: 6 fields ×
N imported pmids) rather than in the schema tests, because long-form
storage is a labeling-module decision and could change without breaking
the schema.

Also covers the three failure-mode contracts the CLI in Batch 4 will
rely on:

* Bad-enum rows surface in ``errors`` and are absent from the parquet.
* Re-importing the same workbook is idempotent (no row duplication).
* ``sample_size`` rejects 0 but accepts blank cells as ``None``.
* The exported workbook actually carries ``DataValidation`` entries on
  the rater-fill columns (a structural guarantee, not just visual).
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from scifield.epistemic.labeling import (
    LABELS_HEADER,
    RATER_FILL_COLS,
    export_to_xlsx,
    import_from_xlsx,
)


def _write_sample(path: Path, n: int = 5) -> pd.DataFrame:
    """Write a synthetic stratified-sample parquet at ``path``."""
    rows = []
    for i in range(n):
        rows.append(
            {
                "pmid": 1000 + i,
                "journal": f"Journal {i % 2}",
                "year": 2010 + i,
                "era": "2010-2019",
                "topic_id": i % 3,
                "title": f"Synthetic title {i}",
                "abstract": f"Synthetic abstract body {i}. Lorem ipsum dolor sit amet.",
            }
        )
    df = pd.DataFrame(rows)
    df.to_parquet(path, index=False)
    return df


def _fill_row(ws, row_idx: int, values: dict) -> None:
    """Write a dict of {field: value} into the rater-fill cells of ``row_idx``."""
    for field, value in values.items():
        col_idx = LABELS_HEADER.index(field) + 1
        ws.cell(row=row_idx, column=col_idx, value=value)


def _good_label_values() -> dict:
    """A complete, valid set of rater-fill values."""
    return {
        "study_design": "RCT",
        "sample_size": 240,
        "has_control": "TRUE",
        "effect_direction": "positive",
        "statistical_claim_present": "TRUE",
        "coi_disclosed_in_abstract": "FALSE",
    }


def test_round_trip_export_fill_import(tmp_path: Path) -> None:
    """Round-trip: 5-row sample -> export -> fill 3 rows -> import.

    Two rows are deliberately left blank to confirm they're skipped
    (not imported as nulls), and the imported parquet must contain
    3 papers × 6 fields = 18 long-form rows.
    """
    sample_path = tmp_path / "sample.parquet"
    xlsx_path = tmp_path / "labels_test.xlsx"
    parquet_out = tmp_path / "handlabel.parquet"

    _write_sample(sample_path, n=5)
    export_to_xlsx(sample_path, xlsx_path, rater_name="test")

    wb = load_workbook(xlsx_path)
    ws = wb["Labels"]
    _fill_row(ws, 2, _good_label_values())
    _fill_row(ws, 3, _good_label_values())
    _fill_row(ws, 4, _good_label_values())
    # rows 5 and 6: leave blank.
    wb.save(xlsx_path)

    summary = import_from_xlsx(xlsx_path, rater_name="test", parquet_out=parquet_out)

    assert summary["n_errors"] == 0, summary["errors"]
    assert summary["n_imported"] == 3
    df = pd.read_parquet(parquet_out)
    assert len(df) == 18
    assert set(df["field"].unique()) == set(RATER_FILL_COLS)
    assert (df["rater"] == "test").all()
    # Each pmid contributes exactly 6 rows.
    assert (df.groupby("pmid").size() == 6).all()


def test_bad_enum_row_surfaces_in_errors(tmp_path: Path) -> None:
    """A bogus ``study_design`` value must error and NOT appear in parquet."""
    sample_path = tmp_path / "sample.parquet"
    xlsx_path = tmp_path / "labels_bad.xlsx"
    parquet_out = tmp_path / "handlabel.parquet"

    _write_sample(sample_path, n=3)
    export_to_xlsx(sample_path, xlsx_path, rater_name="test")

    wb = load_workbook(xlsx_path)
    ws = wb["Labels"]
    _fill_row(ws, 2, _good_label_values())  # row 1 is good.
    bad = _good_label_values()
    bad["study_design"] = "experimental"  # not in the enum
    _fill_row(ws, 3, bad)
    _fill_row(ws, 4, _good_label_values())  # row 3 is good.
    wb.save(xlsx_path)

    summary = import_from_xlsx(xlsx_path, rater_name="test", parquet_out=parquet_out)

    assert summary["n_errors"] == 1
    assert summary["n_imported"] == 2
    err = summary["errors"][0]
    assert err["row"] == 3
    assert err["pmid"] == 1001

    df = pd.read_parquet(parquet_out)
    assert 1001 not in df["pmid"].unique()
    assert set(df["pmid"].unique()) == {1000, 1002}


def test_idempotent_reimport(tmp_path: Path) -> None:
    """Importing the same workbook twice yields the same parquet row count."""
    sample_path = tmp_path / "sample.parquet"
    xlsx_path = tmp_path / "labels_idem.xlsx"
    parquet_out = tmp_path / "handlabel.parquet"

    _write_sample(sample_path, n=3)
    export_to_xlsx(sample_path, xlsx_path, rater_name="alice")

    wb = load_workbook(xlsx_path)
    ws = wb["Labels"]
    for row_idx in (2, 3, 4):
        _fill_row(ws, row_idx, _good_label_values())
    wb.save(xlsx_path)

    s1 = import_from_xlsx(xlsx_path, rater_name="alice", parquet_out=parquet_out)
    df1 = pd.read_parquet(parquet_out)
    n1 = len(df1)
    assert s1["n_errors"] == 0
    assert n1 == 18  # 3 papers x 6 fields

    s2 = import_from_xlsx(xlsx_path, rater_name="alice", parquet_out=parquet_out)
    df2 = pd.read_parquet(parquet_out)
    assert s2["n_errors"] == 0
    assert len(df2) == n1

    # Per-batch (pmid, rater, field) tuples are unique within one import.
    counts = df2.groupby(["pmid", "rater", "field"]).size()
    assert (counts == 1).all()


def test_idempotent_does_not_cross_raters(tmp_path: Path) -> None:
    """Re-import for rater B must NOT drop rater A's rows."""
    sample_path = tmp_path / "sample.parquet"
    xlsx_a = tmp_path / "labels_a.xlsx"
    xlsx_b = tmp_path / "labels_b.xlsx"
    parquet_out = tmp_path / "handlabel.parquet"

    _write_sample(sample_path, n=2)
    export_to_xlsx(sample_path, xlsx_a, rater_name="alice")
    export_to_xlsx(sample_path, xlsx_b, rater_name="bob")

    for path in (xlsx_a, xlsx_b):
        wb = load_workbook(path)
        ws = wb["Labels"]
        _fill_row(ws, 2, _good_label_values())
        _fill_row(ws, 3, _good_label_values())
        wb.save(path)

    import_from_xlsx(xlsx_a, rater_name="alice", parquet_out=parquet_out)
    import_from_xlsx(xlsx_b, rater_name="bob", parquet_out=parquet_out)

    df = pd.read_parquet(parquet_out)
    assert set(df["rater"].unique()) == {"alice", "bob"}
    # 2 raters x 2 papers x 6 fields.
    assert len(df) == 24


def test_sample_size_blank_imports_as_none(tmp_path: Path) -> None:
    """Blank ``sample_size`` cell imports as None (stored as null in parquet)."""
    sample_path = tmp_path / "sample.parquet"
    xlsx_path = tmp_path / "labels_blank.xlsx"
    parquet_out = tmp_path / "handlabel.parquet"

    _write_sample(sample_path, n=1)
    export_to_xlsx(sample_path, xlsx_path, rater_name="test")

    wb = load_workbook(xlsx_path)
    ws = wb["Labels"]
    values = _good_label_values()
    values["study_design"] = "review"
    values["sample_size"] = None  # blank cell
    values["has_control"] = None  # blank cell
    values["effect_direction"] = "na"
    _fill_row(ws, 2, values)
    wb.save(xlsx_path)

    summary = import_from_xlsx(xlsx_path, rater_name="test", parquet_out=parquet_out)
    assert summary["n_errors"] == 0
    assert summary["n_imported"] == 1

    df = pd.read_parquet(parquet_out)
    n_row = df[df["field"] == "sample_size"].iloc[0]
    ctrl_row = df[df["field"] == "has_control"].iloc[0]
    assert n_row["value"] is None or pd.isna(n_row["value"])
    assert ctrl_row["value"] is None or pd.isna(ctrl_row["value"])


def test_sample_size_zero_rejected(tmp_path: Path) -> None:
    """``sample_size=0`` raises a row-level error (Pydantic validator)."""
    sample_path = tmp_path / "sample.parquet"
    xlsx_path = tmp_path / "labels_zero.xlsx"
    parquet_out = tmp_path / "handlabel.parquet"

    _write_sample(sample_path, n=1)
    export_to_xlsx(sample_path, xlsx_path, rater_name="test")

    wb = load_workbook(xlsx_path)
    ws = wb["Labels"]
    values = _good_label_values()
    values["sample_size"] = 0
    _fill_row(ws, 2, values)
    wb.save(xlsx_path)

    summary = import_from_xlsx(xlsx_path, rater_name="test", parquet_out=parquet_out)
    assert summary["n_errors"] == 1
    assert summary["n_imported"] == 0
    err = summary["errors"][0]["error"]
    assert "sample_size" in err.lower() or "1" in err


def test_data_validations_present(tmp_path: Path) -> None:
    """Exported workbook carries ``DataValidation`` on the Labels sheet.

    Structural assertion: at least one DV exists, and the
    ``study_design`` DV formula contains the literal ``RCT`` so we know
    the right column was bound to the right enum.
    """
    sample_path = tmp_path / "sample.parquet"
    xlsx_path = tmp_path / "labels_dv.xlsx"

    _write_sample(sample_path, n=3)
    export_to_xlsx(sample_path, xlsx_path, rater_name="test")

    wb = load_workbook(xlsx_path)
    ws = wb["Labels"]
    dvs = ws.data_validations.dataValidation
    assert len(dvs) >= 1

    # At least one DV mentions RCT in its formula (the study_design one).
    formulas = [dv.formula1 or "" for dv in dvs]
    assert any("RCT" in f for f in formulas), formulas


def test_export_rejects_sample_missing_columns(tmp_path: Path) -> None:
    """Missing helper columns surface a clear error rather than a cryptic crash."""
    sample_path = tmp_path / "sample.parquet"
    xlsx_path = tmp_path / "labels_bad.xlsx"

    bad = pd.DataFrame({"pmid": [1, 2], "journal": ["J", "J"]})  # missing year/title/abstract
    bad.to_parquet(sample_path, index=False)

    with pytest.raises(ValueError, match="missing required columns"):
        export_to_xlsx(sample_path, xlsx_path, rater_name="test")
