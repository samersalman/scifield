"""Tests for V1-S08 arbitration plumbing (`scifield.epistemic.arbitrate`).

Plants a small synthetic two-rater long-form parquet (mirroring what
:func:`scifield.epistemic.labeling.import_from_xlsx` would have written),
drives the three public entry points end-to-end, and asserts:

* Disagreement detection covers all the asymmetry kinds the arbitrator
  needs to see — value-vs-value, value-vs-null, null-vs-value, and the
  one-rater-only pmid that must NOT show up.
* The exported arbitration workbook has the structural surface the
  arbitrator (and downstream import) relies on: correct header,
  correct sheet names, and DataValidation entries on the ``final``
  column that include the right enum strings.
* The round-trip lands in a wide-form parquet whose rows construct
  valid :class:`EpistemicLabel`s, agreement and arbitration sources
  are tagged correctly, and bad arbitration rows surface in ``errors``
  without poisoning the good rows.

Fixture style mirrors :mod:`tests.test_epistemic_labeling` — build a
small DF in-memory, write it to a tmp_path parquet, drive the public
API.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from scifield.epistemic.arbitrate import (
    export_arbitration_xlsx,
    find_disagreements,
    import_arbitration_xlsx,
    load_two_raters,
)
from scifield.epistemic.labeling import RATER_FILL_COLS

RATER_A = "samer"
RATER_B = "partner"


def _good_a_row(pmid: int) -> list[dict]:
    """Build 6 long-form rows for rater A — a 'good' RCT-positive label."""
    base = {
        "study_design": "RCT",
        "sample_size": "240",
        "has_control": "true",
        "effect_direction": "positive",
        "statistical_claim_present": "true",
        "coi_disclosed_in_abstract": "false",
    }
    return [
        {
            "pmid": pmid,
            "rater": RATER_A,
            "field": field,
            "value": value,
            "imported_at": "2026-05-23T00:00:00+00:00",
        }
        for field, value in base.items()
    ]


def _good_b_row(pmid: int) -> list[dict]:
    """Build 6 long-form rows for rater B that AGREE with `_good_a_row`."""
    return [{**r, "rater": RATER_B} for r in _good_a_row(pmid)]


def _build_long_form(tmp_path: Path) -> Path:
    """Build the long-form parquet for fixture (a) and return its path.

    Layout (10-pmid-ish synthetic fixture):

    * pmid 1000, 1001 — full agreement, both raters.
    * pmid 1002 — disagreement on ``study_design`` (RCT vs cohort).
    * pmid 1003 — disagreement on ``sample_size`` (42 vs 60).
    * pmid 1004 — rater_a fills ``effect_direction``=positive,
                  rater_b leaves it blank (null-vs-value).
    * pmid 1005 — rater_a leaves ``has_control`` blank,
                  rater_b fills TRUE (value-vs-null reversed).
    * pmid 1006 — rater_a only — no rater_b rows AT ALL (skipped).
    """
    rows: list[dict] = []

    # pmid 1000, 1001 — full agreement.
    for pmid in (1000, 1001):
        rows.extend(_good_a_row(pmid))
        rows.extend(_good_b_row(pmid))

    # pmid 1002 — study_design disagreement.
    rows.extend(_good_a_row(1002))
    b_1002 = _good_b_row(1002)
    for r in b_1002:
        if r["field"] == "study_design":
            r["value"] = "cohort"
    rows.extend(b_1002)

    # pmid 1003 — sample_size disagreement.
    a_1003 = _good_a_row(1003)
    for r in a_1003:
        if r["field"] == "sample_size":
            r["value"] = "42"
    b_1003 = _good_b_row(1003)
    for r in b_1003:
        if r["field"] == "sample_size":
            r["value"] = "60"
    rows.extend(a_1003)
    rows.extend(b_1003)

    # pmid 1004 — rater_a labels effect_direction; rater_b leaves it blank
    # (drop that field row for rater_b altogether).
    rows.extend(_good_a_row(1004))
    b_1004 = [r for r in _good_b_row(1004) if r["field"] != "effect_direction"]
    rows.extend(b_1004)

    # pmid 1005 — rater_a leaves has_control blank; rater_b fills TRUE.
    # (drop has_control from rater_a's set entirely.)
    a_1005 = [r for r in _good_a_row(1005) if r["field"] != "has_control"]
    rows.extend(a_1005)
    rows.extend(_good_b_row(1005))

    # pmid 1006 — rater_a only.
    rows.extend(_good_a_row(1006))

    long_df = pd.DataFrame(rows, columns=["pmid", "rater", "field", "value", "imported_at"])
    out = tmp_path / "epistemic_handlabel.parquet"
    long_df.to_parquet(out, index=False)
    return out


def test_find_disagreements_detects_all_disagreement_kinds(tmp_path: Path) -> None:
    """Synthetic 6-pmid fixture: 2 agreements, 4 disagreements, 1 single-rater skip."""
    parquet_path = _build_long_form(tmp_path)
    long_df = load_two_raters(parquet_path, RATER_A, RATER_B)

    diffs = find_disagreements(long_df, RATER_A, RATER_B)

    # 4 expected disagreements; pmid 1006 must NOT appear.
    assert len(diffs) == 4, diffs
    assert 1006 not in set(diffs["pmid"].unique())

    # Convert to a tuple-set for order-agnostic comparison.
    triples = {(int(r.pmid), r.field, r.value_a, r.value_b) for r in diffs.itertuples(index=False)}
    assert triples == {
        (1002, "study_design", "RCT", "cohort"),
        (1003, "sample_size", "42", "60"),
        (1004, "effect_direction", "positive", None),
        (1005, "has_control", None, "true"),
    }


def test_export_arbitration_xlsx_has_correct_dropdowns(tmp_path: Path) -> None:
    """Exported workbook carries correct header + enum DVs on `final` column."""
    parquet_path = _build_long_form(tmp_path)
    long_df = load_two_raters(parquet_path, RATER_A, RATER_B)
    diffs = find_disagreements(long_df, RATER_A, RATER_B)
    xlsx_path = tmp_path / "arbitration.xlsx"
    export_arbitration_xlsx(diffs, xlsx_path, RATER_A, RATER_B)

    wb = load_workbook(xlsx_path)
    assert {"Instructions", "Arbitration"}.issubset(set(wb.sheetnames))

    ws = wb["Arbitration"]
    header = tuple(c.value for c in ws[1])
    assert header == ("pmid", "field", "value_a", "value_b", "final")

    # Collect all DV formula1 strings on the Arbitration sheet.
    formulas = [dv.formula1 or "" for dv in ws.data_validations.dataValidation]
    joined = " ".join(formulas)
    # study_design enum (presence of 'RCT' and 'cohort'), BOOL enum
    # (TRUE/FALSE on the has_control disagreement), and effect_direction
    # enum ('positive') should all appear.
    assert "RCT" in joined, formulas
    assert "cohort" in joined, formulas
    assert "TRUE" in joined and "FALSE" in joined, formulas
    assert "positive" in joined, formulas


def test_import_arbitration_xlsx_round_trip(tmp_path: Path) -> None:
    """End-to-end: long-form → export → fill `final` → import → wide parquet."""
    parquet_path = _build_long_form(tmp_path)
    long_df = load_two_raters(parquet_path, RATER_A, RATER_B)
    diffs = find_disagreements(long_df, RATER_A, RATER_B)
    xlsx_path = tmp_path / "arbitration.xlsx"
    export_arbitration_xlsx(diffs, xlsx_path, RATER_A, RATER_B)

    # Fill in the 'final' column. We have to look up each row's field
    # because the workbook's sort order is (field, pmid).
    wb = load_workbook(xlsx_path)
    ws = wb["Arbitration"]
    # Build a (pmid, field) -> chosen final value lookup.
    chosen: dict[tuple[int, str], object] = {
        (1002, "study_design"): "RCT",
        (1003, "sample_size"): 42,
        (1004, "effect_direction"): "positive",
        (1005, "has_control"): "TRUE",
    }
    header_row = [c.value for c in ws[1]]
    final_col_idx = header_row.index("final") + 1
    pmid_col_idx = header_row.index("pmid") + 1
    field_col_idx = header_row.index("field") + 1
    for row in ws.iter_rows(min_row=2):
        pmid_cell = row[pmid_col_idx - 1].value
        field_cell = row[field_col_idx - 1].value
        if pmid_cell is None or field_cell is None:
            continue
        key = (int(pmid_cell), str(field_cell))
        if key in chosen:
            ws.cell(row=row[0].row, column=final_col_idx, value=chosen[key])
    wb.save(xlsx_path)

    out_path = tmp_path / "epistemic_handlabel_final.parquet"
    summary = import_arbitration_xlsx(parquet_path, xlsx_path, RATER_A, RATER_B, out_path)

    assert summary["n_errors"] == 0, summary["errors"]
    assert out_path.exists()
    assert summary["out_path"] == str(out_path)

    df = pd.read_parquet(out_path)
    # PMIDs 1000-1005 are in both raters; 1006 must not appear.
    assert set(df["pmid"].unique()) == {1000, 1001, 1002, 1003, 1004, 1005}

    # Agreement pmids — both raters agreed on everything.
    for agreed_pmid in (1000, 1001):
        row = df[df["pmid"] == agreed_pmid].iloc[0]
        assert row["arbitration_source"] == "agreed"
        assert row["study_design"] == "RCT"
        assert row["sample_size"] == 240
        assert bool(row["has_control"]) is True
        assert row["effect_direction"] == "positive"
        assert bool(row["statistical_claim_present"]) is True
        assert bool(row["coi_disclosed_in_abstract"]) is False

    # 1002 — study_design arbitrated to RCT.
    r1002 = df[df["pmid"] == 1002].iloc[0]
    assert r1002["arbitration_source"] == "arbitrated"
    assert r1002["study_design"] == "RCT"

    # 1003 — sample_size arbitrated to 42.
    r1003 = df[df["pmid"] == 1003].iloc[0]
    assert r1003["arbitration_source"] == "arbitrated"
    assert int(r1003["sample_size"]) == 42

    # 1004 — effect_direction = positive (only rater A had it; B was blank).
    r1004 = df[df["pmid"] == 1004].iloc[0]
    assert r1004["arbitration_source"] == "arbitrated"
    assert r1004["effect_direction"] == "positive"

    # 1005 — has_control = True (only rater B had it; A was blank).
    r1005 = df[df["pmid"] == 1005].iloc[0]
    assert r1005["arbitration_source"] == "arbitrated"
    assert bool(r1005["has_control"]) is True


def test_import_skips_invalid_arbitration_rows(tmp_path: Path) -> None:
    """Bad `final` values surface as errors; good rows still land in parquet."""
    parquet_path = _build_long_form(tmp_path)
    long_df = load_two_raters(parquet_path, RATER_A, RATER_B)
    diffs = find_disagreements(long_df, RATER_A, RATER_B)
    xlsx_path = tmp_path / "arbitration_bad.xlsx"
    export_arbitration_xlsx(diffs, xlsx_path, RATER_A, RATER_B)

    wb = load_workbook(xlsx_path)
    ws = wb["Arbitration"]
    header_row = [c.value for c in ws[1]]
    final_col_idx = header_row.index("final") + 1
    pmid_col_idx = header_row.index("pmid") + 1
    field_col_idx = header_row.index("field") + 1

    # Deliberately bad arbitration: study_design = 'bogus' for 1002,
    # sample_size = 0 for 1003 (Pydantic rejects). 1004 and 1005 get
    # valid finals so they still pass.
    chosen: dict[tuple[int, str], object] = {
        (1002, "study_design"): "bogus",
        (1003, "sample_size"): 0,
        (1004, "effect_direction"): "positive",
        (1005, "has_control"): "TRUE",
    }
    for row in ws.iter_rows(min_row=2):
        pmid_cell = row[pmid_col_idx - 1].value
        field_cell = row[field_col_idx - 1].value
        if pmid_cell is None or field_cell is None:
            continue
        key = (int(pmid_cell), str(field_cell))
        if key in chosen:
            ws.cell(row=row[0].row, column=final_col_idx, value=chosen[key])
    wb.save(xlsx_path)

    out_path = tmp_path / "epistemic_handlabel_final.parquet"
    summary = import_arbitration_xlsx(parquet_path, xlsx_path, RATER_A, RATER_B, out_path)

    # Two errors expected (one Pydantic-enum rejection + one validator
    # rejection for sample_size=0). Agreement and the two good-arb pmids
    # still land.
    assert summary["n_errors"] >= 2, summary["errors"]
    err_pmids = {e.get("pmid") for e in summary["errors"]}
    assert 1002 in err_pmids
    assert 1003 in err_pmids

    df = pd.read_parquet(out_path)
    landed = set(df["pmid"].unique())
    # 1002 and 1003 must NOT be in the parquet (bad arbitration).
    assert 1002 not in landed
    assert 1003 not in landed
    # 1000, 1001 (agreed) + 1004, 1005 (good arbitration) ARE present.
    assert {1000, 1001, 1004, 1005}.issubset(landed)


def test_load_two_raters_errors_on_missing_rater(tmp_path: Path) -> None:
    """Sanity check on `load_two_raters` — a typo'd rater name fails fast."""
    parquet_path = _build_long_form(tmp_path)
    with pytest.raises(ValueError, match="zero rows"):
        load_two_raters(parquet_path, RATER_A, "nobody_by_that_name")


def test_find_disagreements_field_sort_uses_canonical_order(tmp_path: Path) -> None:
    """Disagreements are sorted by (pmid, field) using RATER_FILL_COLS order."""
    parquet_path = _build_long_form(tmp_path)
    long_df = load_two_raters(parquet_path, RATER_A, RATER_B)
    diffs = find_disagreements(long_df, RATER_A, RATER_B)
    # PMIDs must come back ascending.
    assert list(diffs["pmid"]) == sorted(diffs["pmid"].tolist())
    # Each field appears at most once here, but field order must match
    # the canonical RATER_FILL_COLS index whenever there's a tie.
    field_ranks = {f: i for i, f in enumerate(RATER_FILL_COLS)}
    # Build (pmid, rank) and confirm monotonic.
    pairs = [(int(p), field_ranks[f]) for p, f in zip(diffs["pmid"], diffs["field"], strict=False)]
    assert pairs == sorted(pairs)
