"""Excel labeling workflow for V1-S07 epistemic-quality labeling sprint.

The co-rater is (likely) a non-coder, so the hand-labeling tool is a plain
``.xlsx`` workbook rather than a Streamlit app — round-tripped through two
library entry points:

* :func:`export_to_xlsx` — read a stratified sample parquet (written by
  :func:`scifield.epistemic.sampling.stratified_sample`) and emit a
  per-rater workbook with two sheets:

  - **Instructions** — short operational definitions for each field plus
    the active :data:`LABEL_SCHEMA_VERSION` so the rater can verify they
    are on the right schema.
  - **Labels** — one row per abstract; first 5 columns pre-filled and
    locked-by-convention (``pmid, journal, year, title, abstract``);
    last 6 columns are the rater-fill cells with
    :class:`openpyxl.worksheet.datavalidation.DataValidation` enforcing
    the allowed enum values. Blank cells are valid (they mean *not
    filled in yet* and import as ``None``).

* :func:`import_from_xlsx` — read a filled-in workbook, validate every
  row through :class:`scifield.epistemic.schema.EpistemicLabel`, collect
  row-level errors (bad enum, sample_size < 1, bad boolean string, …),
  and append a **long-form** parquet — one row per
  (pmid, rater, field) tuple. Long form is intentional: it lets us
  diff raters per-field without column-name collisions and lets V1-S08
  arbitration produce its own ``rater="arbitrated"`` rows in the same
  parquet without schema migration.

Idempotency: re-importing a workbook for the same ``(pmid, rater)``
overwrites the prior rows for that pair. Implementation drops any
``(pmid, rater)`` tuples in the current batch from the existing parquet
before appending; this means a partial re-label (some pmids missing
from the second import) correctly removes the prior rows for the
re-imported pmids only.

Both functions are library-only (no ``print``, no ``typer.echo``) —
Batch 4 (CLI) is responsible for human-readable output.
"""

from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import ValidationError

from scifield.epistemic.schema import LABEL_SCHEMA_VERSION, EpistemicLabel

__all__ = ["export_to_xlsx", "import_from_xlsx"]


# Column order on the Labels sheet — used by both export and import.
# Helper columns (pre-filled from the sample) come first; rater-fill
# columns (validated on import) come last and match the EpistemicLabel
# field order.
HELPER_COLS: tuple[str, ...] = ("pmid", "journal", "year", "title", "abstract")
RATER_FILL_COLS: tuple[str, ...] = (
    "study_design",
    "sample_size",
    "has_control",
    "effect_direction",
    "statistical_claim_present",
    "coi_disclosed_in_abstract",
)
LABELS_HEADER: tuple[str, ...] = HELPER_COLS + RATER_FILL_COLS

# Closed enums mirrored verbatim from schema.py — duplicated here as
# string literals because openpyxl's DataValidation formula1 wants a
# raw comma-joined string wrapped in double quotes, not a Python tuple.
_STUDY_DESIGN_OPTS = "RCT,cohort,case_control,case_series,review,other"
_EFFECT_DIRECTION_OPTS = "positive,null,negative,mixed,na"
_BOOL_OPTS = "TRUE,FALSE"

# Operational-definition snippets for the Instructions sheet — kept
# terse on purpose; the full pre-registration text lives in
# docs/preregistrations/PR1_epistemic_extraction.md.
_FIELD_DEFINITIONS: tuple[tuple[str, str, str], ...] = (
    (
        "study_design",
        _STUDY_DESIGN_OPTS,
        "Primary design of the work described. Pick the closest match; "
        "use 'other' for methods/protocols/editorials.",
    ),
    (
        "sample_size",
        "integer >= 1, or blank if not stated",
        "Total N reported in the abstract. Leave blank if the abstract "
        "does not state a number; do NOT guess from context.",
    ),
    (
        "has_control",
        _BOOL_OPTS + ", or blank",
        "TRUE if the study uses a control/comparison arm; FALSE if it "
        "explicitly does not; blank if not applicable (e.g., reviews).",
    ),
    (
        "effect_direction",
        _EFFECT_DIRECTION_OPTS,
        "Direction of the primary reported effect. Use 'na' for reviews "
        "or methods papers with no reportable direction.",
    ),
    (
        "statistical_claim_present",
        _BOOL_OPTS,
        "TRUE if the abstract makes any statistical claim (p-value, CI, significance language).",
    ),
    (
        "coi_disclosed_in_abstract",
        _BOOL_OPTS,
        "TRUE if any conflict-of-interest statement appears in the abstract body itself.",
    ),
)


def _write_instructions_sheet(wb: Workbook, rater_name: str) -> None:
    """Populate the Instructions sheet with rater-facing guidance."""
    ws = wb.active
    ws.title = "Instructions"

    header_font = Font(bold=True, size=14)
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    ws["A1"] = f"Epistemic labeling — instructions for rater: {rater_name}"
    ws["A1"].font = header_font

    ws["A3"] = "Schema version:"
    ws["A3"].font = bold
    ws["B3"] = LABEL_SCHEMA_VERSION

    ws["A4"] = "Exported:"
    ws["A4"].font = bold
    ws["B4"] = datetime.now(UTC).date().isoformat()

    ws["A6"] = (
        "Fill in the rater-fill columns on the 'Labels' sheet. Dropdown "
        "validation enforces the allowed values. Blank cells are "
        "permitted (they import as 'not filled in')."
    )
    ws["A6"].alignment = wrap

    # Field definitions table.
    ws["A8"] = "Field"
    ws["B8"] = "Allowed values"
    ws["C8"] = "Definition"
    for col_idx in range(1, 4):
        ws.cell(row=8, column=col_idx).font = bold

    for row_offset, (field, allowed, defn) in enumerate(_FIELD_DEFINITIONS, start=9):
        ws.cell(row=row_offset, column=1, value=field)
        ws.cell(row=row_offset, column=2, value=allowed)
        cell = ws.cell(row=row_offset, column=3, value=defn)
        cell.alignment = wrap

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 80


def _attach_data_validations(ws: Any, n_rows: int) -> None:
    """Attach the 6 rater-fill column validations to rows 2..n_rows+1.

    Each ``DataValidation`` instance is appended to the sheet once, then
    a single A1-style range string is added to it. The ranges target
    rows 2..(n_rows + 1) which is the data block (header is row 1).
    """
    last_row = n_rows + 1  # +1 because row 1 is the header.

    # study_design — list of enum strings.
    dv_design = DataValidation(
        type="list",
        formula1=f'"{_STUDY_DESIGN_OPTS}"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid study_design",
        error=f"Allowed: {_STUDY_DESIGN_OPTS}",
    )
    ws.add_data_validation(dv_design)
    col = get_column_letter(LABELS_HEADER.index("study_design") + 1)
    dv_design.add(f"{col}2:{col}{last_row}")

    # sample_size — positive integers or blank.
    dv_n = DataValidation(
        type="whole",
        operator="greaterThanOrEqual",
        formula1="1",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid sample_size",
        error="sample_size must be an integer >= 1 or blank.",
    )
    ws.add_data_validation(dv_n)
    col = get_column_letter(LABELS_HEADER.index("sample_size") + 1)
    dv_n.add(f"{col}2:{col}{last_row}")

    # has_control — boolean dropdown (allow blank for N/A on reviews).
    dv_ctrl = DataValidation(
        type="list",
        formula1=f'"{_BOOL_OPTS}"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid has_control",
        error=f"Allowed: {_BOOL_OPTS} or blank.",
    )
    ws.add_data_validation(dv_ctrl)
    col = get_column_letter(LABELS_HEADER.index("has_control") + 1)
    dv_ctrl.add(f"{col}2:{col}{last_row}")

    # effect_direction — enum dropdown.
    dv_eff = DataValidation(
        type="list",
        formula1=f'"{_EFFECT_DIRECTION_OPTS}"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid effect_direction",
        error=f"Allowed: {_EFFECT_DIRECTION_OPTS}",
    )
    ws.add_data_validation(dv_eff)
    col = get_column_letter(LABELS_HEADER.index("effect_direction") + 1)
    dv_eff.add(f"{col}2:{col}{last_row}")

    # statistical_claim_present — boolean dropdown.
    dv_stat = DataValidation(
        type="list",
        formula1=f'"{_BOOL_OPTS}"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid statistical_claim_present",
        error=f"Allowed: {_BOOL_OPTS}",
    )
    ws.add_data_validation(dv_stat)
    col = get_column_letter(LABELS_HEADER.index("statistical_claim_present") + 1)
    dv_stat.add(f"{col}2:{col}{last_row}")

    # coi_disclosed_in_abstract — boolean dropdown.
    dv_coi = DataValidation(
        type="list",
        formula1=f'"{_BOOL_OPTS}"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid coi_disclosed_in_abstract",
        error=f"Allowed: {_BOOL_OPTS}",
    )
    ws.add_data_validation(dv_coi)
    col = get_column_letter(LABELS_HEADER.index("coi_disclosed_in_abstract") + 1)
    dv_coi.add(f"{col}2:{col}{last_row}")


def export_to_xlsx(sample_path: Path, out_path: Path, rater_name: str) -> Path:
    """Build a per-rater labeling workbook from a stratified-sample parquet.

    Parameters
    ----------
    sample_path:
        Path to the parquet produced by
        :func:`scifield.epistemic.sampling.stratified_sample`. Must
        contain at least the columns ``pmid, journal, year, title,
        abstract``; extra columns (``era``, ``topic_id``) are ignored.
    out_path:
        Destination ``.xlsx`` path. Parent directory is created if
        missing. Existing files are overwritten.
    rater_name:
        String tag for the rater this workbook is meant for. Stored in
        the Instructions sheet header (also threaded through to the
        import-time parquet).

    Returns
    -------
    Path
        ``out_path`` (echoed back for ergonomics; same as input).
    """
    sample_path = Path(sample_path)
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    df = pd.read_parquet(sample_path)
    missing = [c for c in HELPER_COLS if c not in df.columns]
    if missing:
        raise ValueError(f"sample parquet at {sample_path} missing required columns: {missing}")

    wb = Workbook()
    _write_instructions_sheet(wb, rater_name)

    ws_labels = wb.create_sheet("Labels")
    ws_labels.append(list(LABELS_HEADER))

    # Header styling — bold + light grey fill.
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    header_font = Font(bold=True)
    for col_idx in range(1, len(LABELS_HEADER) + 1):
        cell = ws_labels.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    # Pre-filled helper columns; rater-fill columns left empty.
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in df.itertuples(index=False):
        row_dict = row._asdict()
        ws_labels.append(
            [
                row_dict.get("pmid"),
                row_dict.get("journal"),
                row_dict.get("year"),
                row_dict.get("title"),
                row_dict.get("abstract"),
                None,  # study_design
                None,  # sample_size
                None,  # has_control
                None,  # effect_direction
                None,  # statistical_claim_present
                None,  # coi_disclosed_in_abstract
            ]
        )

    # Column widths — abstract gets the wide column with wrap.
    width_by_name = {
        "pmid": 12,
        "journal": 22,
        "year": 8,
        "title": 50,
        "abstract": 80,
        "study_design": 16,
        "sample_size": 14,
        "has_control": 14,
        "effect_direction": 18,
        "statistical_claim_present": 26,
        "coi_disclosed_in_abstract": 28,
    }
    for idx, name in enumerate(LABELS_HEADER, start=1):
        ws_labels.column_dimensions[get_column_letter(idx)].width = width_by_name[name]

    # Apply wrap to the abstract column on the data rows.
    abstract_col_idx = LABELS_HEADER.index("abstract") + 1
    for row_idx in range(2, len(df) + 2):
        ws_labels.cell(row=row_idx, column=abstract_col_idx).alignment = wrap

    if len(df) > 0:
        _attach_data_validations(ws_labels, n_rows=len(df))

    wb.save(out_path)
    return out_path


def _coerce_cell(field: str, raw: Any) -> Any:
    """Convert a raw openpyxl cell value to the type expected by the schema.

    Returns the coerced value on success, or raises :class:`ValueError`
    with a human-readable message that the caller turns into a row-level
    error entry.

    Coercion rules:

    * Blank cells (``None`` or ``""``) → ``None`` for every field.
    * The 3 boolean fields accept Python booleans, or the strings
      ``"TRUE"`` / ``"FALSE"`` (case-insensitive). Anything else raises.
    * ``sample_size`` accepts ints and float-valued ints (e.g. openpyxl
      sometimes returns ``240.0``); anything else raises. The Pydantic
      validator then enforces the ``>= 1`` rule.
    * Enum fields are passed through as strings; Pydantic rejects bad
      values at construction time.
    """
    if raw is None or (isinstance(raw, str) and raw.strip() == ""):
        return None

    if field in ("has_control", "statistical_claim_present", "coi_disclosed_in_abstract"):
        if isinstance(raw, bool):
            return raw
        if isinstance(raw, str):
            up = raw.strip().upper()
            if up == "TRUE":
                return True
            if up == "FALSE":
                return False
            raise ValueError(f"{field}: expected TRUE/FALSE/blank, got {raw!r}")
        raise ValueError(f"{field}: expected TRUE/FALSE/blank, got {raw!r}")

    if field == "sample_size":
        if isinstance(raw, bool):  # bool is a subclass of int — reject explicitly.
            raise ValueError(f"sample_size: expected integer or blank, got {raw!r}")
        if isinstance(raw, int):
            return raw
        if isinstance(raw, float):
            if raw.is_integer():
                return int(raw)
            raise ValueError(f"sample_size: expected integer or blank, got {raw!r}")
        if isinstance(raw, str):
            try:
                return int(raw.strip())
            except ValueError as exc:
                raise ValueError(f"sample_size: expected integer or blank, got {raw!r}") from exc
        raise ValueError(f"sample_size: expected integer or blank, got {raw!r}")

    # Enum strings — let Pydantic do the membership check.
    if isinstance(raw, str):
        return raw.strip()
    raise ValueError(f"{field}: expected string enum value, got {raw!r}")


def import_from_xlsx(xlsx_path: Path, rater_name: str, parquet_out: Path) -> dict:
    """Import a filled-in labeling workbook into the long-form labels parquet.

    Parameters
    ----------
    xlsx_path:
        Path to the filled-in workbook. Must have a ``Labels`` sheet
        with the canonical 11-column header.
    rater_name:
        Tag identifying the human (or aggregator) producing these
        labels. Stored verbatim in the ``rater`` column of every
        produced parquet row.
    parquet_out:
        Destination parquet path. If it already exists, prior rows
        whose ``(pmid, rater)`` matches any pmid in this batch are
        dropped before appending — making re-imports idempotent for
        the same rater.

    Returns
    -------
    dict
        Summary with keys ``n_rows`` (total Labels rows seen, excluding
        header), ``n_imported`` (rows that passed validation),
        ``n_errors`` (rows that failed), ``errors`` (list of
        ``{row, pmid, error}`` dicts), and ``out_path`` (string form of
        ``parquet_out``).
    """
    xlsx_path = Path(xlsx_path)
    parquet_out = Path(parquet_out)
    parquet_out.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(xlsx_path, data_only=True)
    if "Labels" not in wb.sheetnames:
        raise ValueError(f"workbook {xlsx_path} has no 'Labels' sheet")
    ws = wb["Labels"]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration as exc:
        raise ValueError(f"workbook {xlsx_path} 'Labels' sheet is empty") from exc

    header = tuple(h for h in header if h is not None)
    if header != LABELS_HEADER:
        raise ValueError(
            f"workbook {xlsx_path} 'Labels' header mismatch: expected {LABELS_HEADER}, got {header}"
        )

    imported_at = datetime.now(UTC).isoformat()
    errors: list[dict] = []
    new_records: list[dict] = []
    seen_pmids: set[int] = set()
    n_rows = 0

    for row_idx, row in enumerate(rows_iter, start=2):
        # Skip wholly-empty rows (openpyxl emits trailing Nones for
        # workbooks that have been touched past the data block).
        if row is None or all(cell is None or cell == "" for cell in row):
            continue
        n_rows += 1

        # Pad short rows out to the header width — openpyxl truncates
        # trailing Nones in some edge cases.
        padded = list(row) + [None] * (len(LABELS_HEADER) - len(row))
        record = dict(zip(LABELS_HEADER, padded[: len(LABELS_HEADER)], strict=False))

        pmid_raw = record.get("pmid")
        try:
            pmid = int(pmid_raw) if pmid_raw is not None else None
        except (TypeError, ValueError):
            errors.append(
                {"row": row_idx, "pmid": pmid_raw, "error": f"invalid pmid: {pmid_raw!r}"}
            )
            continue
        if pmid is None:
            errors.append({"row": row_idx, "pmid": None, "error": "missing pmid"})
            continue

        # Coerce + validate the 6 rater-fill fields.
        coerced: dict[str, Any] = {}
        coerce_error: str | None = None
        for field in RATER_FILL_COLS:
            try:
                coerced[field] = _coerce_cell(field, record.get(field))
            except ValueError as exc:
                coerce_error = str(exc)
                break

        if coerce_error is not None:
            errors.append({"row": row_idx, "pmid": pmid, "error": coerce_error})
            continue

        # Rows where every rater-fill cell is blank are "not yet
        # labeled" — silently skip rather than flagging missing-required
        # errors. Required fields with a None value mean a rater started
        # but did not finish the row; we still flag those (Pydantic will
        # complain).
        if all(coerced[f] is None for f in RATER_FILL_COLS):
            continue

        try:
            label = EpistemicLabel(**coerced)
        except ValidationError as exc:
            errors.append({"row": row_idx, "pmid": pmid, "error": exc.errors()[0]["msg"]})
            continue

        seen_pmids.add(pmid)
        label_dict = label.model_dump()
        for field in RATER_FILL_COLS:
            value = label_dict[field]
            # Normalize to a single string column type for the long-form
            # parquet: booleans -> "true"/"false", ints -> str, None
            # -> None, enum strings pass through as-is.
            if value is None:
                stored: str | None = None
            elif isinstance(value, bool):
                stored = "true" if value else "false"
            elif isinstance(value, int):
                stored = str(value)
            else:
                stored = str(value)
            new_records.append(
                {
                    "pmid": pmid,
                    "rater": rater_name,
                    "field": field,
                    "value": stored,
                    "imported_at": imported_at,
                }
            )

    new_df = pd.DataFrame(
        new_records,
        columns=["pmid", "rater", "field", "value", "imported_at"],
    )

    if parquet_out.exists():
        existing = pd.read_parquet(parquet_out)
        if not existing.empty and seen_pmids:
            mask_drop = (existing["rater"] == rater_name) & (existing["pmid"].isin(seen_pmids))
            existing = existing[~mask_drop]
        combined = pd.concat([existing, new_df], ignore_index=True)
    else:
        combined = new_df

    combined.to_parquet(parquet_out, index=False)

    return {
        "n_rows": n_rows,
        "n_imported": len(seen_pmids),
        "n_errors": len(errors),
        "errors": errors,
        "out_path": str(parquet_out),
    }
