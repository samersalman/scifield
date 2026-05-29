"""Arbitration plumbing for V1-S08 inter-rater label reconciliation.

V1-S07 left us with one long-form parquet
(`data/v1/epistemic_handlabel.parquet`) where two human raters
(Samer + partner) each contribute one row per (pmid, rater, field, value,
imported_at) tuple via :func:`scifield.epistemic.labeling.import_from_xlsx`.

V1-S08's arbitration step needs to turn that into a single
wide-form parquet at `data/v1/epistemic_handlabel_final.parquet` —
one row per pmid, one column per :class:`EpistemicLabel` field, plus an
``arbitration_source`` column that says whether the value came directly
from rater agreement or from a tie-break by the arbitrator. That final
parquet is what V1-S09 (kappa + LLM-vs-arbitrated agreement) consumes.

The hand-off between the rater-pair stage and the final wide parquet is
a *3-column arbitration workbook* (`value_a`, `value_b`, `final`) — only
the rows where the two raters disagree end up in the workbook. The
arbitrator (Samer) opens it, picks the winning value field-by-field, and
saves. :func:`import_arbitration_xlsx` then reads back the filled-in
workbook AND the agreement set from the original long-form parquet and
emits the wide-form final parquet.

Design choices worth flagging:

* **Long → wide pivot only at the very last step.** The long-form parquet
  is the source of truth all the way through; we pivot rater-to-columns
  only when we need a side-by-side diff (this module's job). Long form
  scales painlessly to N>2 raters if we ever decide to recruit more, and
  it keeps the ``import_from_xlsx`` idempotency story clean.

* **Null-vs-value counts as disagreement.** If rater A labels a field
  and rater B leaves it blank for the same pmid, that's a meaningful
  disagreement (B might genuinely think the field is N/A; A might have
  over-labeled). The arbitrator decides whether the blank wins or the
  value wins. PMIDs that one rater hasn't touched at all (zero rows
  from that rater for that pmid) are SKIPPED — they're not "yet
  labeled" and shouldn't appear in the arbitration workbook.

* **Dropdowns mirror :mod:`scifield.epistemic.labeling` exactly.** The
  ``final`` column reuses the same enum strings via direct import — no
  duplication of ``"RCT,cohort,..."`` lives in this file. If the
  upstream enum set changes, the arbitration workbook follows.

* **Validation on import.** Every constructed :class:`EpistemicLabel`
  must pass Pydantic validation. Failures become row-level errors but
  do NOT abort the write — we mirror the
  :func:`scifield.epistemic.labeling.import_from_xlsx` pattern so an
  arbitrator typo on one row doesn't lose the other 199 good rows.

Library-only — no prints, no typer. CLI wrapper (Wave 2) is responsible
for human-readable output.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import ValidationError

from scifield.epistemic.labeling import (
    _BOOL_OPTS,
    _EFFECT_DIRECTION_OPTS,
    _STUDY_DESIGN_OPTS,
    RATER_FILL_COLS,
)
from scifield.epistemic.schema import LABEL_SCHEMA_VERSION, EpistemicLabel

__all__ = [
    "load_two_raters",
    "find_disagreements",
    "export_arbitration_xlsx",
    "import_arbitration_xlsx",
]


# Header for the arbitration workbook 'Arbitration' sheet.
_ARBITRATION_HEADER: tuple[str, ...] = ("pmid", "field", "value_a", "value_b", "final")

# Enum-string lookup per field, keyed off the same field names exported
# by :mod:`scifield.epistemic.labeling`. Re-used by the 'final' column's
# DataValidation entries in :func:`export_arbitration_xlsx`.
_FIELD_TO_OPTS: dict[str, str] = {
    "study_design": _STUDY_DESIGN_OPTS,
    "has_control": _BOOL_OPTS,
    "effect_direction": _EFFECT_DIRECTION_OPTS,
    "statistical_claim_present": _BOOL_OPTS,
    "coi_disclosed_in_abstract": _BOOL_OPTS,
}

# Canonical sort order for the `field` column — mirrors RATER_FILL_COLS.
_FIELD_ORDER: dict[str, int] = {name: idx for idx, name in enumerate(RATER_FILL_COLS)}

# Final wide-form parquet schema. Explicit so an empty arbitration run
# still writes a typed parquet that downstream V1-S09 can open.
_FINAL_PARQUET_SCHEMA = pa.schema(
    [
        ("pmid", pa.int64()),
        ("study_design", pa.string()),
        ("sample_size", pa.int64()),
        ("has_control", pa.bool_()),
        ("effect_direction", pa.string()),
        ("statistical_claim_present", pa.bool_()),
        ("coi_disclosed_in_abstract", pa.bool_()),
        ("arbitration_source", pa.string()),
    ]
)


def load_two_raters(
    handlabel_parquet: Path,
    rater_a: str,
    rater_b: str,
) -> pd.DataFrame:
    """Load the long-form handlabel parquet restricted to two raters.

    Parameters
    ----------
    handlabel_parquet:
        Path to the long-form parquet written by
        :func:`scifield.epistemic.labeling.import_from_xlsx`. Must have
        columns ``pmid, rater, field, value, imported_at``.
    rater_a, rater_b:
        Rater tags as they appear in the ``rater`` column. Order is
        significant only insofar as downstream
        :func:`find_disagreements` reports ``value_a`` / ``value_b``
        following the same order.

    Returns
    -------
    pd.DataFrame
        Long-form DF with columns ``[pmid, rater, field, value]``
        restricted to the two raters. Both raters must have at least
        one row in the parquet — otherwise the caller has named a rater
        that hasn't labeled anything yet, which is almost always a
        typo and we'd rather fail fast.

    Raises
    ------
    FileNotFoundError
        If ``handlabel_parquet`` does not exist.
    ValueError
        If either rater has zero rows in the parquet.
    """
    handlabel_parquet = Path(handlabel_parquet)
    if not handlabel_parquet.exists():
        raise FileNotFoundError(f"handlabel parquet not found: {handlabel_parquet}")

    df = pd.read_parquet(handlabel_parquet)
    required_cols = {"pmid", "rater", "field", "value"}
    missing = required_cols - set(df.columns)
    if missing:
        raise ValueError(
            f"handlabel parquet {handlabel_parquet} missing columns: {sorted(missing)}"
        )

    sub = df[df["rater"].isin([rater_a, rater_b])][["pmid", "rater", "field", "value"]].copy()
    n_a = int((sub["rater"] == rater_a).sum())
    n_b = int((sub["rater"] == rater_b).sum())
    if n_a == 0:
        raise ValueError(f"rater {rater_a!r} has zero rows in {handlabel_parquet}")
    if n_b == 0:
        raise ValueError(f"rater {rater_b!r} has zero rows in {handlabel_parquet}")
    return sub.reset_index(drop=True)


def _pivot_wide(long_df: pd.DataFrame, rater_a: str, rater_b: str) -> pd.DataFrame:
    """Pivot the long-form DF into one row per (pmid, field) with cols a/b.

    Result columns: ``[pmid, field, value_a, value_b]`` where missing
    (rater, field) combinations are ``None``. Used internally by
    :func:`find_disagreements`.
    """
    # Aggregate any accidental duplicate (pmid, rater, field) rows by
    # taking the last value — the handlabel parquet's idempotency
    # guarantee means duplicates should not occur in practice, but we
    # defend against the corner case so groupby doesn't silently raise.
    deduped = long_df.drop_duplicates(subset=["pmid", "rater", "field"], keep="last")

    wide = deduped.pivot_table(
        index=["pmid", "field"],
        columns="rater",
        values="value",
        aggfunc="first",
    ).reset_index()
    wide.columns.name = None

    # Guarantee both rater columns exist even when one rater never
    # labeled a particular field (pivot_table drops missing columns).
    if rater_a not in wide.columns:
        wide[rater_a] = None
    if rater_b not in wide.columns:
        wide[rater_b] = None
    wide = wide.rename(columns={rater_a: "value_a", rater_b: "value_b"})
    return wide[["pmid", "field", "value_a", "value_b"]]


def find_disagreements(
    long_df: pd.DataFrame,
    rater_a: str,
    rater_b: str,
) -> pd.DataFrame:
    """Return one row per (pmid, field) where the two raters disagree.

    A disagreement counts when:

    * Both raters labeled the field and gave different values.
    * One rater labeled the field and the other left it blank (null vs.
      value asymmetry).

    PMIDs where one rater has ZERO rows across all fields (i.e. that
    rater simply hasn't labeled that pmid yet) are skipped entirely —
    they would otherwise flood the workbook with spurious disagreements
    just because the partner is behind on labeling.

    Sort order: ``pmid`` ascending, then ``field`` in the canonical
    :data:`scifield.epistemic.labeling.RATER_FILL_COLS` order.

    Parameters
    ----------
    long_df:
        Long-form DF with at least ``[pmid, rater, field, value]``
        columns. Typically the output of :func:`load_two_raters`.
    rater_a, rater_b:
        The two rater tags. ``value_a`` in the result corresponds to
        ``rater_a``; ``value_b`` to ``rater_b``.

    Returns
    -------
    pd.DataFrame
        Columns ``[pmid, field, value_a, value_b]``, one row per
        disagreement.
    """
    # Restrict to the two raters and filter out pmids only one rater
    # has touched at all.
    sub = long_df[long_df["rater"].isin([rater_a, rater_b])]
    if sub.empty:
        return pd.DataFrame(columns=["pmid", "field", "value_a", "value_b"])

    pmids_a = set(sub.loc[sub["rater"] == rater_a, "pmid"].unique())
    pmids_b = set(sub.loc[sub["rater"] == rater_b, "pmid"].unique())
    both = pmids_a & pmids_b
    sub = sub[sub["pmid"].isin(both)]
    if sub.empty:
        return pd.DataFrame(columns=["pmid", "field", "value_a", "value_b"])

    wide = _pivot_wide(sub, rater_a, rater_b)

    # Treat NaN cells as None so the value/null comparison below is
    # unambiguous regardless of pandas dtype quirks. We rebuild the
    # columns as object-dtype lists rather than using .map(), because
    # pandas will sometimes re-introduce NaN when an object column
    # has a mix of strings and Nones.
    def _norm(v: Any) -> Any:
        if v is None:
            return None
        if isinstance(v, float) and pd.isna(v):
            return None
        return v

    # Build raw Python lists so None survives instead of being coerced
    # back to NaN by pandas' object-column quirks.
    norm_a = [_norm(v) for v in wide["value_a"].tolist()]
    norm_b = [_norm(v) for v in wide["value_b"].tolist()]
    pmid_list = wide["pmid"].tolist()
    field_list = wide["field"].tolist()

    diff_records = [
        {
            "pmid": pmid_list[i],
            "field": field_list[i],
            "value_a": norm_a[i],
            "value_b": norm_b[i],
        }
        for i in range(len(pmid_list))
        if norm_a[i] != norm_b[i]
    ]
    diff_records.sort(key=lambda r: (r["pmid"], _FIELD_ORDER.get(r["field"], len(_FIELD_ORDER))))
    if not diff_records:
        return pd.DataFrame(columns=["pmid", "field", "value_a", "value_b"])

    # Build the DF with explicit object-dtype Series so None doesn't get
    # converted to NaN. Going through pd.DataFrame(list-of-dicts) does
    # implicit type inference that converts None → NaN whenever a column
    # has any non-None entry.
    diff = pd.DataFrame(
        {
            "pmid": pd.Series([r["pmid"] for r in diff_records], dtype="int64"),
            "field": pd.Series([r["field"] for r in diff_records], dtype=object),
            "value_a": pd.Series([r["value_a"] for r in diff_records], dtype=object),
            "value_b": pd.Series([r["value_b"] for r in diff_records], dtype=object),
        }
    )
    # Defensive: replace any lingering NaN with None on object columns
    # so downstream comparisons (and tests) see a clean None.
    for col in ("value_a", "value_b"):
        diff[col] = diff[col].where(diff[col].notna(), None)
    return diff


def _write_arbitration_instructions(wb: Workbook, rater_a: str, rater_b: str) -> None:
    """Populate the 'Instructions' sheet for the arbitration workbook."""
    ws = wb.active
    ws.title = "Instructions"

    header_font = Font(bold=True, size=14)
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    ws["A1"] = "Epistemic-label arbitration"
    ws["A1"].font = header_font

    ws["A3"] = "Schema version:"
    ws["A3"].font = bold
    ws["B3"] = LABEL_SCHEMA_VERSION

    ws["A4"] = "Rater A:"
    ws["A4"].font = bold
    ws["B4"] = rater_a

    ws["A5"] = "Rater B:"
    ws["A5"].font = bold
    ws["B5"] = rater_b

    ws["A7"] = (
        "Go to the 'Arbitration' sheet. Each row is one (pmid, field) "
        "where rater A and rater B gave different labels (including "
        "one-blank cases). Pick the winning value in the 'final' column "
        "using the dropdown. Blank 'final' = not yet arbitrated and will "
        "be skipped on import."
    )
    ws["A7"].alignment = wrap

    ws["A9"] = (
        "For null-vs-value rows: if you believe the blank is correct, "
        "type the literal word 'blank' (lowercase) into 'final' to lock "
        "that decision in. Anything truly empty will be treated as 'not "
        "yet arbitrated'."
    )
    ws["A9"].alignment = wrap

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60


def export_arbitration_xlsx(
    disagreements: pd.DataFrame,
    out_path: Path,
    rater_a: str,
    rater_b: str,
) -> Path:
    """Write a 5-column arbitration workbook the arbitrator fills in.

    The 'Arbitration' sheet has header
    ``[pmid, field, value_a, value_b, final]`` with one row per
    disagreement. The 'final' column carries field-specific
    :class:`openpyxl.worksheet.datavalidation.DataValidation` dropdowns
    keyed off the row's ``field`` value (e.g. study_design rows get the
    study_design enum; boolean fields get TRUE/FALSE; sample_size has
    no enum so no DV is attached — the arbitrator types a number).

    Implementation note: rows are sorted by (field, pmid) within the
    workbook so each field's DataValidation can cover one contiguous
    cell range. This deviates from :func:`find_disagreements`'s
    (pmid, field)-sorted output — the export ordering is local to the
    workbook only, and the input DF is not mutated.

    Parameters
    ----------
    disagreements:
        Output of :func:`find_disagreements`.
    out_path:
        Destination .xlsx path. Parent dir is created if missing.
    rater_a, rater_b:
        Rater tags, stored in the Instructions sheet so the arbitrator
        knows which column is which.

    Returns
    -------
    Path
        ``out_path`` (echoed back for ergonomics).
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    _write_arbitration_instructions(wb, rater_a, rater_b)

    ws = wb.create_sheet("Arbitration")
    ws.append(list(_ARBITRATION_HEADER))

    header_fill = PatternFill("solid", fgColor="DDDDDD")
    header_font = Font(bold=True)
    for col_idx in range(1, len(_ARBITRATION_HEADER) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    # Sort the export rows by (field, pmid) so each field's DV
    # range is contiguous. The input DF is left untouched.
    export_df = disagreements.copy()
    if len(export_df) > 0:
        export_df["_field_rank"] = (
            export_df["field"].map(_FIELD_ORDER).fillna(len(_FIELD_ORDER)).astype(int)
        )
        export_df = export_df.sort_values(by=["_field_rank", "pmid"], kind="stable").drop(
            columns="_field_rank"
        )
    export_df = export_df.reset_index(drop=True)

    for row_tuple in export_df.itertuples(index=False):
        ws.append(
            [
                row_tuple.pmid,
                row_tuple.field,
                row_tuple.value_a,
                row_tuple.value_b,
                None,  # final — arbitrator fills.
            ]
        )

    # Column widths.
    width_by_name = {
        "pmid": 12,
        "field": 26,
        "value_a": 16,
        "value_b": 16,
        "final": 18,
    }
    for idx, name in enumerate(_ARBITRATION_HEADER, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width_by_name[name]

    # Build per-field contiguous blocks for DataValidation. Sort gives
    # us blocks of identical field values; we walk the export DF and
    # attach one DV per block.
    final_col_letter = get_column_letter(_ARBITRATION_HEADER.index("final") + 1)
    if len(export_df) > 0:
        cur_field: str | None = None
        block_start: int | None = None
        # Excel rows are 1-indexed; +2 because row 1 is the header and
        # the iloc index starts at 0.
        for i, field in enumerate(export_df["field"].tolist() + [None]):  # sentinel
            xl_row = i + 2
            if field != cur_field:
                # Close the prior block (if any) and start a new one.
                if cur_field is not None and block_start is not None:
                    _attach_final_validation(
                        ws,
                        field=cur_field,
                        col_letter=final_col_letter,
                        first_row=block_start,
                        last_row=xl_row - 1,
                    )
                cur_field = field
                block_start = xl_row

    wb.save(out_path)
    return out_path


def _attach_final_validation(
    ws: Any,
    *,
    field: str,
    col_letter: str,
    first_row: int,
    last_row: int,
) -> None:
    """Attach the right enum DV to ``final`` cells for one field block.

    For ``sample_size`` no enum DV exists (free-text integer) — caller
    just gets no DV on that block, which is the intended behavior.
    For null-vs-value disagreements the arbitrator may also write the
    literal word ``"blank"``, so all enum DVs are constructed with
    ``allow_blank=True`` and we DO NOT set ``showErrorMessage=True``
    on the field — that would block the literal ``blank`` keyword.
    """
    if field not in _FIELD_TO_OPTS:
        # sample_size and any future free-text fields: no list DV.
        return
    opts = _FIELD_TO_OPTS[field]
    # We deliberately leave showErrorMessage False so the arbitrator
    # can type the literal word "blank" to record a null-wins decision.
    dv = DataValidation(
        type="list",
        formula1=f'"{opts},blank"',
        allow_blank=True,
        showErrorMessage=False,
    )
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first_row}:{col_letter}{last_row}")


def _coerce_final_value(field: str, raw: Any) -> Any | None:
    """Convert a raw 'final' cell to the type expected by EpistemicLabel.

    Returns:
    * ``None`` for blanks (the row is "not yet arbitrated", caller skips).
    * ``None`` for the literal sentinel ``"blank"`` (the arbitrator
      explicitly chose null-wins).
    * Coerced bool / int / str otherwise.

    Raises ``ValueError`` with a human-readable message on coercion
    failures — caller turns this into a row-level error entry.

    Distinguishing "not yet arbitrated" (skip) from "explicit blank"
    (use None) is the reason for the ``"blank"`` literal: a totally
    empty cell could mean either, and we'd rather make the arbitrator
    say it out loud.
    """
    if raw is None or (isinstance(raw, str) and raw.strip() == ""):
        return _NOT_ARBITRATED

    if isinstance(raw, str) and raw.strip().lower() == "blank":
        return None

    if field in ("has_control", "statistical_claim_present", "coi_disclosed_in_abstract"):
        if isinstance(raw, bool):
            return raw
        if isinstance(raw, str):
            up = raw.strip().upper()
            if up == "TRUE":
                return True
            if up == "FALSE":
                return False
            raise ValueError(f"{field}: expected TRUE/FALSE/blank, got {raw!r}")
        raise ValueError(f"{field}: expected TRUE/FALSE/blank, got {raw!r}")

    if field == "sample_size":
        if isinstance(raw, bool):
            raise ValueError(f"sample_size: expected integer or blank, got {raw!r}")
        if isinstance(raw, int):
            return raw
        if isinstance(raw, float):
            if raw.is_integer():
                return int(raw)
            raise ValueError(f"sample_size: expected integer or blank, got {raw!r}")
        if isinstance(raw, str):
            try:
                return int(raw.strip())
            except ValueError as exc:
                raise ValueError(f"sample_size: expected integer or blank, got {raw!r}") from exc
        raise ValueError(f"sample_size: expected integer or blank, got {raw!r}")

    if isinstance(raw, str):
        return raw.strip()
    raise ValueError(f"{field}: expected string enum value, got {raw!r}")


# Sentinel so callers can distinguish "blank cell, skip this row" from
# "arbitrator chose null". Module-private; never crosses the public API.
_NOT_ARBITRATED = object()


def _coerce_agreement_value(field: str, raw: str | None) -> Any:
    """Convert a string from the long-form parquet back to typed values.

    Mirror image of the value-serialization in
    :func:`scifield.epistemic.labeling.import_from_xlsx` (which stores
    everything as a string for parquet column-type uniformity).

    Raises ``ValueError`` on malformed input — but in practice this
    should never happen because the values were written through the
    same coercion path on import.
    """
    if raw is None:
        return None
    if isinstance(raw, float) and pd.isna(raw):
        return None
    if not isinstance(raw, str):
        # Defense: if upstream ever stops stringifying, surface the bug
        # rather than silently misclassifying.
        raise ValueError(f"{field}: expected stored str, got {type(raw).__name__}")
    s = raw.strip()
    if field in ("has_control", "statistical_claim_present", "coi_disclosed_in_abstract"):
        if s.lower() == "true":
            return True
        if s.lower() == "false":
            return False
        raise ValueError(f"{field}: expected 'true'/'false', got {raw!r}")
    if field == "sample_size":
        try:
            return int(s)
        except ValueError as exc:
            raise ValueError(f"sample_size: expected int string, got {raw!r}") from exc
    # Enum string — pass through.
    return s


def import_arbitration_xlsx(
    handlabel_parquet: Path,
    arbitration_xlsx: Path,
    rater_a: str,
    rater_b: str,
    out_path: Path,
) -> dict:
    """Combine agreements + arbitrated values into the final wide parquet.

    Workflow:

    1. Re-load the long-form handlabel parquet, restricted to the two
       raters. For each (pmid, field) where the two raters agree (same
       non-null value), record the agreed value.
    2. Open the arbitration workbook's 'Arbitration' sheet. Each row
       provides ``final`` = the arbitrator's tie-break value.
    3. Group both streams by ``pmid``; build an :class:`EpistemicLabel`
       per pmid. PMIDs touched only by agreement OR only by arbitration
       (i.e. partial coverage of the 6 fields) are silently allowed —
       Pydantic enforces required fields and raises if essential ones
       are missing; that surfaces as a row-level error.
    4. Write the wide-form parquet.

    Parameters
    ----------
    handlabel_parquet:
        Source long-form parquet (V1-S07 import output).
    arbitration_xlsx:
        Filled-in arbitration workbook produced by
        :func:`export_arbitration_xlsx`.
    rater_a, rater_b:
        Same tags used in the original export.
    out_path:
        Destination parquet path.

    Returns
    -------
    dict
        ``{"n_pmids": int, "n_agreed_fields": int,
        "n_arbitrated_fields": int, "n_errors": int,
        "errors": list[dict], "out_path": str}``.
    """
    handlabel_parquet = Path(handlabel_parquet)
    arbitration_xlsx = Path(arbitration_xlsx)
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    long_df = load_two_raters(handlabel_parquet, rater_a, rater_b)

    # ----- Agreements: same value from both raters for the same field.
    # Only include pmids that BOTH raters have at least one row for —
    # mirrors the policy in find_disagreements.
    pmids_a = set(long_df.loc[long_df["rater"] == rater_a, "pmid"].unique())
    pmids_b = set(long_df.loc[long_df["rater"] == rater_b, "pmid"].unique())
    pmids_both = pmids_a & pmids_b
    paired = long_df[long_df["pmid"].isin(pmids_both)]

    wide = _pivot_wide(paired, rater_a, rater_b)

    agreement_rows: dict[int, dict[str, Any]] = {}
    n_agreed_fields = 0
    errors: list[dict] = []
    for _, row in wide.iterrows():
        va = row["value_a"]
        vb = row["value_b"]
        if isinstance(va, float) and pd.isna(va):
            va = None
        if isinstance(vb, float) and pd.isna(vb):
            vb = None
        if va is None and vb is None:
            # Neither rater filled this field — nothing to record.
            continue
        if va != vb:
            # Disagreement — handled via the arbitration workbook.
            continue
        try:
            coerced = _coerce_agreement_value(row["field"], va)
        except ValueError as exc:
            errors.append(
                {
                    "pmid": int(row["pmid"]),
                    "field": row["field"],
                    "source": "agreement",
                    "error": str(exc),
                }
            )
            continue
        pmid = int(row["pmid"])
        agreement_rows.setdefault(pmid, {})[row["field"]] = coerced
        n_agreed_fields += 1

    # ----- Arbitrated values from the workbook.
    wb = load_workbook(arbitration_xlsx, data_only=True)
    if "Arbitration" not in wb.sheetnames:
        raise ValueError(f"workbook {arbitration_xlsx} has no 'Arbitration' sheet")
    ws = wb["Arbitration"]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration as exc:
        raise ValueError(f"workbook {arbitration_xlsx} 'Arbitration' sheet is empty") from exc
    header = tuple(h for h in header if h is not None)
    if header != _ARBITRATION_HEADER:
        raise ValueError(
            f"workbook {arbitration_xlsx} 'Arbitration' header mismatch: "
            f"expected {_ARBITRATION_HEADER}, got {header}"
        )

    arbitrated_rows: dict[int, dict[str, Any]] = {}
    n_arbitrated_fields = 0
    for row_idx, raw in enumerate(rows_iter, start=2):
        if raw is None or all(c is None or c == "" for c in raw):
            continue
        padded = list(raw) + [None] * (len(_ARBITRATION_HEADER) - len(raw))
        rec = dict(zip(_ARBITRATION_HEADER, padded[: len(_ARBITRATION_HEADER)], strict=False))

        pmid_raw = rec.get("pmid")
        field = rec.get("field")
        pmid_arb: int | None
        try:
            pmid_arb = int(pmid_raw) if pmid_raw is not None else None
        except (TypeError, ValueError):
            errors.append(
                {
                    "row": row_idx,
                    "pmid": pmid_raw,
                    "field": field,
                    "source": "arbitration",
                    "error": f"invalid pmid: {pmid_raw!r}",
                }
            )
            continue
        if pmid_arb is None or field is None:
            errors.append(
                {
                    "row": row_idx,
                    "pmid": pmid_arb,
                    "field": field,
                    "source": "arbitration",
                    "error": "missing pmid or field",
                }
            )
            continue
        if field not in RATER_FILL_COLS:
            errors.append(
                {
                    "row": row_idx,
                    "pmid": pmid_arb,
                    "field": field,
                    "source": "arbitration",
                    "error": f"unknown field: {field!r}",
                }
            )
            continue

        try:
            coerced = _coerce_final_value(field, rec.get("final"))
        except ValueError as exc:
            errors.append(
                {
                    "row": row_idx,
                    "pmid": pmid_arb,
                    "field": field,
                    "source": "arbitration",
                    "error": str(exc),
                }
            )
            continue

        if coerced is _NOT_ARBITRATED:
            # Empty 'final' cell — the arbitrator hasn't decided yet;
            # leave the field unset and Pydantic will surface a
            # required-field error if applicable.
            continue
        arbitrated_rows.setdefault(pmid_arb, {})[field] = coerced
        n_arbitrated_fields += 1

    # ----- Build EpistemicLabel rows.
    all_pmids = sorted(set(agreement_rows) | set(arbitrated_rows))
    records: list[dict[str, Any]] = []
    for pmid in all_pmids:
        agreed = agreement_rows.get(pmid, {})
        arbitrated = arbitrated_rows.get(pmid, {})
        # Arbitrated values override agreement values for any
        # collisions — in practice they should never overlap (a field
        # is either agreement or disagreement, never both), but if a
        # human re-imports stale data we'd rather respect the most
        # recent decision.
        merged: dict[str, Any] = {**agreed, **arbitrated}

        # Source label per pmid: "agreed" if every field came from
        # agreement, "arbitrated" if anything came from arbitration.
        source = "arbitrated" if arbitrated else "agreed"

        try:
            label = EpistemicLabel(**merged)
        except ValidationError as exc:
            errors.append(
                {
                    "pmid": pmid,
                    "field": None,
                    "source": source,
                    "error": exc.errors()[0]["msg"],
                }
            )
            continue
        record = label.model_dump()
        record["pmid"] = pmid
        record["arbitration_source"] = source
        records.append(record)

    # ----- Write the wide-form parquet with an explicit schema.
    col_to_arr: dict[str, list[Any]] = {name: [] for name in _FINAL_PARQUET_SCHEMA.names}
    for rec in records:
        col_to_arr["pmid"].append(rec["pmid"])
        col_to_arr["study_design"].append(rec.get("study_design"))
        col_to_arr["sample_size"].append(rec.get("sample_size"))
        col_to_arr["has_control"].append(rec.get("has_control"))
        col_to_arr["effect_direction"].append(rec.get("effect_direction"))
        col_to_arr["statistical_claim_present"].append(rec.get("statistical_claim_present"))
        col_to_arr["coi_disclosed_in_abstract"].append(rec.get("coi_disclosed_in_abstract"))
        col_to_arr["arbitration_source"].append(rec["arbitration_source"])

    table = pa.table(col_to_arr, schema=_FINAL_PARQUET_SCHEMA)
    pq.write_table(table, out_path)

    return {
        "n_pmids": len(records),
        "n_agreed_fields": n_agreed_fields,
        "n_arbitrated_fields": n_arbitrated_fields,
        "n_errors": len(errors),
        "errors": errors,
        "out_path": str(out_path),
    }
